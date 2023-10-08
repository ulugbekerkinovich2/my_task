import datetime
import json
import os

import openpyxl
from django.conf import settings
from django.db.models import Count
from django.http import JsonResponse, FileResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from openpyxl.workbook import Workbook

from robots.models import Robot


@csrf_exempt
def robot_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            serial = data.get('serial')
            model = data.get('model')
            version = data.get('version')

            if Robot.objects.filter(model=model).exists():
                raise ValueError('Robot already exists')

            instance_robot = Robot.objects.create(
                model=model, serial=serial, version=version, created=timezone.now()
            )
            instance_robot.save()

            return JsonResponse({'status': 200, 'data': data})

        except ValueError as e:
            return JsonResponse({'status': 400, 'error': str(e)})
    else:
        return JsonResponse({'status': 405, 'error': 'Invalid request method'})


def download_excel(request):
    try:
        last_week = timezone.now() - timezone.timedelta(days=7)

        data = Robot.objects.filter(created__gte=last_week).values('model', 'serial').annotate(count_week=Count('id'))

        wb = Workbook()
        for d in data:
            model = d['model']
            serial = d['serial']
            count_week = d['count_week']

            if model not in wb.sheetnames:
                wb.create_sheet(title=model)
            sheet = wb[model]

            if sheet.max_row == 1:
                sheet.append(['Модель', 'Серийный номер', 'Количество за неделю'])

            sheet.append([model, serial, count_week])

        del wb[wb.sheetnames[0]]

        file_path = os.path.join(settings.MEDIA_ROOT, 'robots.xlsx')

        wb.save(file_path)

        file_url = request.build_absolute_uri(settings.MEDIA_URL + 'robots.xlsx')

        return JsonResponse({"file_url": file_url})

    except Exception as e:
        return JsonResponse({'status': 500, 'error': str(e)})
